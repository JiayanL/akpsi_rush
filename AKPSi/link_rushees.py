import sys
import random
import openpyxl
import PyPDF2


def create_candidate_page_dict(excel_file, pdf_file, k):
    # Read the Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Read the PDF file
    pdf_reader = PyPDF2.PdfFileReader(pdf_file)

    # Initialize the dictionary
    candidate_page_dict = {}

    # Initialize the page number
    page_number = 2

    # Iterate through the rows in the Excel file
    for row in range(2, sheet.max_row + 1):
        # Skip hidden rows
        if sheet.row_dimensions[row].hidden:
            continue

        candidate_name = sheet.cell(row=row, column=1).value

        # Skip the kth page in the PDF
        if page_number == k - 1:
            page_number += 1

        # Add the candidate name and page number to the dictionary
        candidate_page_dict[candidate_name] = page_number

        # Increment the page number
        page_number += 1
    del candidate_page_dict[None]

    return candidate_page_dict


def randomize_candidate_pages(candidate_page_dict, k):
    candidates_up_to_k = []
    candidates_from_k = []

    for candidate, page_number in candidate_page_dict.items():
        if page_number < k - 1:
            candidates_up_to_k.append(candidate)
        elif page_number > k - 1:
            candidates_from_k.append(candidate)

    # Create the page number ranges for the two segments
    page_range_up_to_k = list(range(2, k))
    page_range_from_k = list(range(k + 1, len(candidate_page_dict) + 4))

    # Shuffle the page numbers within their respective ranges
    random.shuffle(page_range_up_to_k)
    random.shuffle(page_range_from_k)

    randomized_candidate_page_dict = {}
    for i, candidate in enumerate(candidates_up_to_k):
        randomized_candidate_page_dict[candidate] = page_range_up_to_k[i]

    for i, candidate in enumerate(candidates_from_k):
        try:
            randomized_candidate_page_dict[candidate] = page_range_from_k[i]
        except:
            print(i)

    return randomized_candidate_page_dict


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print("Usage: python randomize_candidate_pages.py excel_file pdf_file k")
        sys.exit(1)

    excel_file = sys.argv[1]
    pdf_file = sys.argv[2]
    k = int(sys.argv[3])

    candidate_page_dict = create_candidate_page_dict(excel_file, pdf_file, k)
    randomized_candidate_page_dict = randomize_candidate_pages(
        candidate_page_dict, k)

    # Print the randomized candidate page dictionary
    for candidate, page in randomized_candidate_page_dict.items():
        print(f"{candidate}: {page}")
